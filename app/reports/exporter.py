from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook


def export_report_excel(report: dict, output_path: str) -> str:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    for key, value in report.get("statistics", {}).items():
        summary_sheet.append([key, value])

    issues_sheet = workbook.create_sheet("Issues")
    issues_sheet.append(
        [
            "index",
            "audit_group",
            "audit_object",
            "checkpoint",
            "result",
            "issue_description",
            "clause_reference",
            "evidence_source",
            "severity",
            "rectification_suggestion",
            "confidence",
        ]
    )
    for item in report.get("issue_list", []):
        issues_sheet.append(
            [
                item.get("index"),
                item.get("audit_group"),
                item.get("audit_object"),
                item.get("checkpoint"),
                item.get("result"),
                item.get("issue_description"),
                item.get("clause_reference"),
                item.get("evidence_source"),
                item.get("severity"),
                item.get("rectification_suggestion"),
                item.get("confidence"),
            ]
        )

    evidence_sheet = workbook.create_sheet("Evidence")
    evidence_sheet.append(
        [
            "issue_index",
            "source_file_id",
            "source_file_name",
            "page",
            "snippet",
            "field_path",
            "extracted_field_source",
            "locator",
            "rule_id",
            "clause_id",
        ]
    )
    for item in report.get("issue_list", []):
        evidence_chain = item.get("internal", {}).get("evidence_chain", [])
        if not evidence_chain:
            continue
        for evidence in evidence_chain:
            evidence_sheet.append(
                [
                    item.get("index"),
                    evidence.get("source_file_id"),
                    evidence.get("source_file_name"),
                    evidence.get("page"),
                    evidence.get("snippet"),
                    evidence.get("field_path"),
                    evidence.get("extracted_field_source"),
                    json.dumps(evidence.get("locator"), ensure_ascii=False) if evidence.get("locator") else "",
                    evidence.get("rule_id"),
                    evidence.get("clause_id"),
                ]
            )

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return str(target)
